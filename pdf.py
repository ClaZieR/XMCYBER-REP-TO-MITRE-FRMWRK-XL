import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def read_excel(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')
    return df[['ID', 'name', 'tactics', 'platforms']].set_index('ID')

def detect_mitre_techniques(pdf_path):
    mitre_pattern = re.compile(r'T\d{3,}(?:\.\d{1,3})?')
    occurrences = {}

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                found_techniques = mitre_pattern.findall(text)
                for technique in found_techniques:
                    occurrences[technique] = occurrences.get(technique, 0) + 1

    return occurrences

# Paths to your PDF and Excel file
pdf_path = r'D:\VS Projects\PY\Report - CredentialDump.pdf'
excel_path = r'D:\VS Projects\PY\enterprise-attack-v14.1.xlsx'

# Read the Excel data and detect MITRE techniques in the PDF
excel_data = read_excel(excel_path)
technique_occurrences = detect_mitre_techniques(pdf_path)

# Collecting results and finding the first non-decimal technique
first_non_decimal_technique_name = None
results = []

for technique in technique_occurrences.keys():
    if '.' not in technique and first_non_decimal_technique_name is None and technique in excel_data.index:
        first_non_decimal_technique_name = excel_data.at[technique, 'name'].replace(' ', '_').replace('/', '_')
        break  # Stop after finding the first non-decimal technique

for technique in technique_occurrences.keys():
    if '.' in technique and technique in excel_data.index:  # Only consider techniques with a decimal point
        name = excel_data.at[technique, 'name']
        tactic = excel_data.at[technique, 'tactics']
        platform = excel_data.at[technique, 'platforms']
        results.append([technique, name, tactic, platform])

# Determine the output file name
output_file_name = first_non_decimal_technique_name if first_non_decimal_technique_name else "mitre_techniques_results"
output_path = fr'D:\VS Projects\PY\{output_file_name}.xlsx'

# Create a DataFrame and export to Excel
results_df = pd.DataFrame(results, columns=['Technique ID', 'Name', 'Tactics', 'Platform'])
results_df.to_excel(output_path, index=False)

# Adjust column width and row height using openpyxl
wb = load_workbook(output_path)
sheet = wb.active

for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

for row in sheet.iter_rows():
    sheet.row_dimensions[row[0].row].height = 15

wb.save(output_path)
print(f"Results exported to {output_path}")
